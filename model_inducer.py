#! /usr/bin/env python3

import openpyxl
import sys
import pprint

# FIXME: This should come from the command line
START_INDEX = 3

class ModelInducer(object):
    def run(self, args):
        # We want to answer question:
        # Given data set arg,
        # Holding the value for 'group column' steady,
        # Do we identify duplication in 'distinct column'?

        data_set_path = args[0]

        # 1-based index of the column.  FIXME, should be using alphabetical
        # references really.
        group_column = int(args[1])
        distinct_column = int(args[2])

        unique_groups = {
        }

        wb = openpyxl.load_workbook(args[0])
        ws = wb.active

        row_count = ws.max_row
        for i in range(START_INDEX, row_count):
            group_key = ws.cell(row=i, column=group_column).value
            group_value = ws.cell(row=i, column=distinct_column).value
            unique_groups.setdefault(group_key, []).append(group_value)


        for group, all_values in unique_groups.items():
            distinct_values = set(all_values)
            if len(distinct_values) != len(all_values):
                print("denormalization happened on group", group)
            
        

if __name__ == '__main__':
    obj = ModelInducer()
    obj.run(sys.argv[1:])
